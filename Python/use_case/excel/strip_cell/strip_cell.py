"""
读取目标 Excel 文件，去除每个单元格的首尾的空格，并保存到新的 Excel 文件中
保留原始 Excel 的所有格式
"""
from pathlib import Path
from openpyxl import load_workbook

CURRENT_DIR = Path(__file__).resolve().parent

# 读取目标 Excel 文件
src_excel_filepath = CURRENT_DIR / "data/2024年度关注漏洞.xlsx"
wb = load_workbook(src_excel_filepath)
ws = wb.active

# 在原始工作簿上修改单元格值
for row in ws.iter_rows():
    for cell in row:
        # 只对字符串类型的值进行strip操作
        if cell.value is not None and isinstance(cell.value, str):
            cell.value = cell.value.strip()

# 保存为新的 Excel 文件
target_excel_filepath = CURRENT_DIR / "data/2024年度关注漏洞_去除首尾空格.xlsx"
wb.save(target_excel_filepath)

print(f"新的 Excel 文件已保存到 {target_excel_filepath}")